import openpyxl
from openpyxl import Workbook
workbook=Workbook()
sheet=workbook.active#选取当前活跃的工作表
sheet.title='考勤'#表名
data=[
    ['datatime','name','age'],
    ['2021,11,1','xiaobei','21'],
    ['2021,11,2','xiaohui','22']
]
for row in data:
    sheet.append(row)
print('第一个单元格{}'.format(sheet['A1']))
workbook.save('考勤表.xlsx')#保存工作表
#读取工作表
wb=openpyxl.load_workbook('考勤表.xlsx')
print('工作表的名称{}'.format(wb.sheetnames))
sheet=wb['考勤']
for row in sheet.rows:
    for i in row:
        print(i.value)